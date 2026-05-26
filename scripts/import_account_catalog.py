from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from db.engine import get_engine, session_factory
from db.models import AccountCatalog


REQUIRED_HEADERS = {"GRUPO", "SUB GRUPO", "RUBRO", "NATURALEZA"}
VALID_TYPE_SECTION = {
    "activo": {"corriente", "no_corriente"},
    "pasivo": {"corriente", "no_corriente"},
    "patrimonio": {"patrimonio"},
    "ingreso": {"ingresos", "otros_ingresos"},
    "costo": {"costo_ventas"},
    "gasto": {"gastos_operativos", "gastos_financieros", "otros_egresos"},
}


@dataclass(frozen=True)
class CatalogRecord:
    code: str
    name: str
    account_type: str
    section: str
    niif_code: str = ""
    normal_balance: str = ""
    parent_code: str = ""
    aliases: tuple[str, ...] = ()
    display_order: int = 0
    source: str = "niif_pyme"
    required_model_account: bool = False


INTERNAL_ACCOUNTS: list[CatalogRecord] = [
    CatalogRecord("cash", "Efectivo y Equivalentes al Efectivo", "activo", "corriente", "111", "deudora", "niif_11", ("caja", "banco", "bancos", "efectivo y equivalentes de efectivo"), 1110, "niif_pyme_enriched", True),
    CatalogRecord("accounts_receivable", "Deudores Comerciales y otras cuentas por cobrar", "activo", "corriente", "113", "deudora", "niif_11", ("cuentas por cobrar", "cuentas por cobrar clientes"), 1130, "niif_pyme_enriched", True),
    CatalogRecord("inventory", "Inventarios", "activo", "corriente", "115", "deudora", "niif_11", ("inventario",), 1150, "niif_pyme_enriched", True),
    CatalogRecord("ppe_real_estate", "Bienes Inmuebles", "activo", "no_corriente", "121.01", "deudora", "niif_121", ("vivienda", "propiedad inmueble"), 1211, "niif_pyme_enriched", True),
    CatalogRecord("ppe_equipment", "Mobiliario y Equipos", "activo", "no_corriente", "121.02", "deudora", "niif_121", ("equipos", "mobiliario"), 1212, "niif_pyme_enriched", True),
    CatalogRecord("ppe_vehicles", "Vehiculos", "activo", "no_corriente", "121.03", "deudora", "niif_121", ("vehiculos", "vehiculo"), 1213, "niif_pyme_enriched", True),
    CatalogRecord("accum_depreciation", "Depreciacion Acumulada", "activo", "no_corriente", "121.90", "acreedora", "niif_121", ("depreciacion acumulada",), 1219, "niif_pyme_enriched", True),
    CatalogRecord("suppliers", "Cuentas por pagar comerciales", "pasivo", "corriente", "212", "acreedora", "niif_21", ("proveedores",), 2120, "niif_pyme_enriched", True),
    CatalogRecord("credit_cards", "Tarjetas de Credito", "pasivo", "corriente", "213.01", "acreedora", "niif_213", ("tarjetas", "tarjetas de credito"), 2131, "niif_pyme_enriched", True),
    CatalogRecord("taxes_payable", "Impuestos por Pagar", "pasivo", "corriente", "219", "acreedora", "niif_21", ("impuestos por pagar",), 2190, "niif_pyme_enriched", True),
    CatalogRecord("accrued_expenses", "Gastos Devengados por pagar", "pasivo", "corriente", "214", "acreedora", "niif_21", ("gastos acumulados", "gastos acumulados por pagar"), 2140, "niif_pyme_enriched", True),
    CatalogRecord("loans_mortgage", "Creditos Hipotecarios", "pasivo", "no_corriente", "221.01", "acreedora", "niif_221", ("creditos hipotecarios",), 2211, "niif_pyme_enriched", True),
    CatalogRecord("loans_consumo", "Creditos Consumo", "pasivo", "no_corriente", "221.02", "acreedora", "niif_221", ("creditos consumo",), 2212, "niif_pyme_enriched", True),
    CatalogRecord("loans_personal", "Creditos Personales", "pasivo", "no_corriente", "221.03", "acreedora", "niif_221", ("creditos personales",), 2213, "niif_pyme_enriched", True),
    CatalogRecord("loans_pledge", "Creditos Prendarios", "pasivo", "no_corriente", "221.04", "acreedora", "niif_221", ("creditos prendarios",), 2214, "niif_pyme_enriched", True),
    CatalogRecord("loans_commercial", "Creditos Comerciales", "pasivo", "no_corriente", "221.05", "acreedora", "niif_221", ("creditos comerciales",), 2215, "niif_pyme_enriched", True),
    CatalogRecord("capital", "Capital Social", "patrimonio", "patrimonio", "311", "acreedora", "niif_31", ("capital",), 3110, "niif_pyme_enriched", True),
    CatalogRecord("retained_earnings", "Resultados Acumulados", "patrimonio", "patrimonio", "321", "acreedora", "niif_32", ("resultado acumulado", "resultados acumulados"), 3210, "niif_pyme_enriched", True),
    CatalogRecord("current_earnings", "Resultados del Ejercicio", "patrimonio", "patrimonio", "322", "acreedora", "niif_32", ("resultado del ejercicio", "utilidad del ejercicio"), 3220, "niif_pyme_enriched", True),
    CatalogRecord("legal_reserve", "Reserva Legal", "patrimonio", "patrimonio", "325", "acreedora", "niif_32", ("reservas legales",), 3250, "niif_pyme_enriched", False),
    CatalogRecord("revenue", "Ventas", "ingreso", "ingresos", "411", "acreedora", "niif_41", ("ingresos", "ventas"), 4110, "niif_pyme_enriched", True),
    CatalogRecord("cogs", "Costo de los Productos Vendidos", "costo", "costo_ventas", "511", "deudora", "niif_51", ("costo de venta", "costo de ventas"), 5110, "niif_pyme_enriched", True),
    CatalogRecord("operating_expenses", "Gastos Operativos", "gasto", "gastos_operativos", "61", "deudora", "niif_6", ("gastos operativos",), 6100, "niif_pyme_enriched", True),
    CatalogRecord("exp_salaries", "Sueldos y Salarios", "gasto", "gastos_operativos", "611.01", "deudora", "operating_expenses", ("sueldos", "salarios", "sueldos y salarios"), 6111, "niif_pyme_enriched", False),
    CatalogRecord("exp_services", "Servicios Publicos", "gasto", "gastos_operativos", "612.01", "deudora", "operating_expenses", ("servicios publicos", "servicios públicos"), 6121, "niif_pyme_enriched", False),
    CatalogRecord("depreciation_expense", "Gasto por Depreciacion", "gasto", "gastos_operativos", "613.01", "deudora", "operating_expenses", ("depreciaciones", "gasto por depreciacion"), 6131, "niif_pyme_enriched", True),
    CatalogRecord("financial_expenses", "Gastos Financieros", "gasto", "gastos_financieros", "615", "deudora", "operating_expenses", ("gastos financieros",), 6150, "niif_pyme_enriched", True),
    CatalogRecord("exp_alcaldia_dgi", "Alcaldia y DGI", "gasto", "gastos_operativos", "619.01", "deudora", "niif_619", ("alcaldia y dgi", "alcaldía y dgi"), 6191, "niif_pyme_enriched", False),
    CatalogRecord("exp_fuel", "Combustible", "gasto", "gastos_operativos", "619.02", "deudora", "niif_619", ("combustibles",), 6192, "niif_pyme_enriched", False),
    CatalogRecord("exp_advertising", "Publicidad", "gasto", "gastos_operativos", "619.03", "deudora", "niif_619", ("publicidad",), 6193, "niif_pyme_enriched", False),
    CatalogRecord("exp_maintenance", "Mantenimientos", "gasto", "gastos_operativos", "619.04", "deudora", "niif_619", ("mantenimiento",), 6194, "niif_pyme_enriched", False),
    CatalogRecord("exp_rent", "Renta", "gasto", "gastos_operativos", "619.05", "deudora", "niif_619", ("alquiler", "arrendamiento"), 6195, "niif_pyme_enriched", False),
    CatalogRecord("exp_insurance", "Seguros", "gasto", "gastos_operativos", "619.06", "deudora", "niif_619", ("seguro",), 6196, "niif_pyme_enriched", False),
    CatalogRecord("exp_other", "Otros Gastos", "gasto", "gastos_operativos", "619.99", "deudora", "niif_619", ("otros gastos", "otros"), 6199, "niif_pyme_enriched", False),
]


SKIP_NIIF_CODES = {record.niif_code for record in INTERNAL_ACCOUNTS if re.fullmatch(r"\d+", record.niif_code)}


def load_catalog_records(path: Path) -> list[CatalogRecord]:
    rows = _load_rows(path)
    records: dict[str, CatalogRecord] = {}
    for idx, row in enumerate(rows, start=1):
        group_code, group_name = _split_code_name(row.get("GRUPO"))
        subgroup_code, subgroup_name = _split_code_name(row.get("SUB GRUPO"))
        rubro_code, rubro_name = _split_code_name(row.get("RUBRO"))
        nature = _normal_balance(row.get("NATURALEZA"))
        if not group_code or not subgroup_code or not rubro_code:
            continue
        account_type, section = _type_section(group_code, subgroup_code, rubro_code, rubro_name)
        records.setdefault(
            f"niif_{group_code}",
            CatalogRecord(f"niif_{group_code}", group_name, account_type, _group_section(account_type), group_code, nature, "", (), int(group_code) * 100000, "niif_pyme", False),
        )
        if subgroup_code not in {"61"}:
            records.setdefault(
                f"niif_{subgroup_code}",
                CatalogRecord(f"niif_{subgroup_code}", subgroup_name, account_type, section, subgroup_code, nature, f"niif_{group_code}", (), int(subgroup_code) * 1000, "niif_pyme", False),
            )
        if rubro_code not in SKIP_NIIF_CODES:
            records.setdefault(
                f"niif_{rubro_code}",
                CatalogRecord(f"niif_{rubro_code}", rubro_name, account_type, section, rubro_code, nature, f"niif_{subgroup_code}" if subgroup_code != "61" else "operating_expenses", (), int(rubro_code) * 10 + idx, "niif_pyme", False),
            )
    for record in INTERNAL_ACCOUNTS:
        records[record.code] = record
    _validate_records(records.values())
    return sorted(records.values(), key=lambda record: (record.display_order, record.code))


def import_catalog(path: Path, *, apply: bool = False, deactivate_missing: bool = False, session=None) -> dict[str, Any]:
    records = load_catalog_records(path)
    summary = {"records": len(records), "created": 0, "updated": 0, "deactivated": 0, "required": 0}
    summary["required"] = sum(1 for record in records if record.required_model_account)
    if not apply:
        summary["mode"] = "dry-run"
        summary["sample"] = [_record_dict(record) for record in records[:8]]
        return summary

    own_session = session is None
    if own_session:
        factory = session_factory(get_engine())
        session = factory()
    try:
        existing = {account.code: account for account in session.query(AccountCatalog).all()}
        incoming_codes = {record.code for record in records}
        now = datetime.now(timezone.utc)
        for record in records:
            account = existing.get(record.code)
            if account is None:
                account = AccountCatalog(id=record.code, code=record.code, created_at=now)
                session.add(account)
                summary["created"] += 1
            else:
                summary["updated"] += 1
            _apply_record(account, record)
        if deactivate_missing:
            for code, account in existing.items():
                if code not in incoming_codes and account.source in {"system", "niif_pyme", "niif_pyme_enriched"}:
                    account.active = 0
                    summary["deactivated"] += 1
        session.commit()
        summary["mode"] = "apply"
        return summary
    except Exception:
        session.rollback()
        raise
    finally:
        if own_session:
            session.close()


def _load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Catalogo"] if "Catalogo" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(cell or "").strip() for cell in rows[0]]
        _validate_headers(headers)
        return [dict(zip(headers, row)) for row in rows[1:]]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            _validate_headers(reader.fieldnames or [])
            return list(reader)
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            raise ValueError("El JSON del catalogo debe ser una lista de registros.")
        if data:
            _validate_headers(data[0].keys())
        return data
    raise ValueError("Formato no soportado. Use .xlsx, .csv o .json.")


def _validate_headers(headers: Iterable[str]) -> None:
    found = {str(header or "").strip().upper() for header in headers}
    missing = REQUIRED_HEADERS - found
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(sorted(missing))}.")


def _split_code_name(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    match = re.match(r"^(\d+(?:\.\d+)?)\s*(.*)$", text)
    if not match:
        return "", text
    return match.group(1).strip(), match.group(2).strip()


def _normal_balance(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if raw.startswith("deud"):
        return "deudora"
    if raw.startswith("acre"):
        return "acreedora"
    return raw


def _group_section(account_type: str) -> str:
    return {
        "activo": "corriente",
        "pasivo": "corriente",
        "patrimonio": "patrimonio",
        "ingreso": "ingresos",
        "costo": "costo_ventas",
        "gasto": "gastos_operativos",
    }.get(account_type, "corriente")


def _type_section(group_code: str, subgroup_code: str, rubro_code: str, rubro_name: str) -> tuple[str, str]:
    if group_code == "1":
        return "activo", "no_corriente" if subgroup_code.startswith("12") else "corriente"
    if group_code == "2":
        return "pasivo", "no_corriente" if subgroup_code.startswith("22") else "corriente"
    if group_code == "3":
        return "patrimonio", "patrimonio"
    if group_code == "4":
        return "ingreso", "otros_ingresos" if subgroup_code.startswith("42") else "ingresos"
    if group_code == "5":
        return "costo", "costo_ventas"
    if group_code == "6":
        if rubro_code == "615" or "financier" in rubro_name.lower():
            return "gasto", "gastos_financieros"
        if subgroup_code.startswith("62"):
            return "gasto", "otros_egresos"
        return "gasto", "gastos_operativos"
    return "activo", "corriente"


def _validate_records(records: Iterable[CatalogRecord]) -> None:
    seen: set[str] = set()
    required: set[str] = set()
    for record in records:
        if record.code in seen:
            raise ValueError(f"Codigo duplicado en catalogo: {record.code}")
        seen.add(record.code)
        if record.section not in VALID_TYPE_SECTION.get(record.account_type, set()):
            raise ValueError(f"Combinacion tipo/seccion invalida: {record.code} {record.account_type}/{record.section}")
        if record.required_model_account:
            required.add(record.code)
    missing = {record.code for record in INTERNAL_ACCOUNTS if record.required_model_account} - required
    if missing:
        raise ValueError(f"Faltan cuentas obligatorias del modelo: {', '.join(sorted(missing))}")


def _record_dict(record: CatalogRecord) -> dict[str, Any]:
    return {
        "code": record.code,
        "niif_code": record.niif_code,
        "name": record.name,
        "account_type": record.account_type,
        "section": record.section,
        "normal_balance": record.normal_balance,
        "parent_code": record.parent_code,
        "aliases": list(record.aliases),
        "display_order": record.display_order,
        "source": record.source,
        "required_model_account": record.required_model_account,
    }


def _apply_record(account: AccountCatalog, record: CatalogRecord) -> None:
    account.code = record.code
    account.niif_code = record.niif_code or None
    account.name = record.name
    account.account_type = record.account_type
    account.section = record.section
    account.normal_balance = record.normal_balance or None
    account.parent_code = record.parent_code or None
    account.aliases_json = json.dumps(list(record.aliases), ensure_ascii=False)
    account.display_order = int(record.display_order or 0)
    account.required_model_account = 1 if record.required_model_account else 0
    account.source = record.source
    account.active = 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa el catalogo NIIF PYME enriquecido a SQLite.")
    parser.add_argument("--file", required=True, help="Ruta del archivo .xlsx, .csv o .json")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true", help="Valida y muestra resumen sin escribir")
    mode.add_argument("--apply", action="store_true", help="Actualiza account_catalog")
    parser.add_argument("--deactivate-missing", action="store_true", help="Desactiva cuentas base que no esten en el archivo enriquecido")
    args = parser.parse_args()
    summary = import_catalog(Path(args.file), apply=args.apply, deactivate_missing=args.deactivate_missing)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
